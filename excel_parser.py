import re
from datetime import datetime
import openpyxl
from models import Order, OrderItem


def parse_order_number(text):
    """
    Извлекает номер заказа из текста
    Пример: "Заказ покупателя № 2351 от 8 декабря 2025 г." -> "2351"
    """
    if not text:
        return None
    
    # Ищем номер после символа №
    match = re.search(r'№\s*(\d+)', str(text))
    if match:
        return match.group(1)
    
    # Альтернативный поиск просто цифр
    match = re.search(r'(\d+)', str(text))
    if match:
        return match.group(1)
    
    return None


def parse_order_date(text):
    """
    Извлекает дату из текста заказа
    Пример: "Заказ покупателя № 2351 от 8 декабря 2025 г."
    """
    if not text:
        return datetime.now().date()
    
    # Ищем дату в формате "от DD месяц YYYY"
    months = {
        'января': 1, 'февраля': 2, 'марта': 3, 'апреля': 4,
        'мая': 5, 'июня': 6, 'июля': 7, 'августа': 8,
        'сентября': 9, 'октября': 10, 'ноября': 11, 'декабря': 12
    }
    
    match = re.search(r'от\s+(\d+)\s+(\w+)\s+(\d{4})', str(text))
    if match:
        day = int(match.group(1))
        month_name = match.group(2)
        year = int(match.group(3))
        
        month = months.get(month_name.lower(), 1)
        return datetime(year, month, day).date()
    
    # Если не нашли, возвращаем текущую дату
    return datetime.now().date()


def parse_excel_file(filepath, filename):
    """
    Парсит Excel файл заказа из 1С
    
    Структура файла:
    - Строка 3: Заголовок с номером и датой заказа
    - Строка 9: Заголовки колонок
    - Строки 11+: Данные товаров
      - Колонка 1 (индекс 0): № строки
      - Колонка 6 (индекс 5): Наименование товара
      - Колонка 17 (индекс 16): Код товара
      - Колонка 20 (индекс 19): Количество
      - Колонка 23 (индекс 22): Единица измерения
    
    Returns:
        Order: Объект заказа с товарами
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    
    # Извлекаем информацию о заказе из строки 3
    header_cell = ws.cell(row=3, column=2)  # Колонка B, строка 3
    header_text = header_cell.value
    
    order_number = parse_order_number(header_text)
    order_date = parse_order_date(header_text)
    
    if not order_number:
        order_number = f"ORDER_{datetime.now().strftime('%Y%m%d%H%M%S')}"
    
    # Создаем заказ
    order = Order(
        order_number=order_number,
        order_date=order_date,
        filename=filename,
        status='новый'
    )
    
    # Парсим товары начиная со строки 11
    items = []
    row_num = 11
    
    while row_num <= ws.max_row:
        # Проверяем есть ли номер строки (колонка 1)
        row_number_cell = ws.cell(row=row_num, column=2)  # Колонка B
        if not row_number_cell.value:
            row_num += 1
            continue
        
        # Извлекаем данные товара
        name_cell = ws.cell(row=row_num, column=7)  # Колонка G (наименование)
        quantity_cell = ws.cell(row=row_num, column=21)  # Колонка U (количество)
        unit_cell = ws.cell(row=row_num, column=24)  # Колонка X (единица)
        code_cell = ws.cell(row=row_num, column=18)  # Колонка R (код)
        
        name = name_cell.value
        quantity = quantity_cell.value
        
        # Пропускаем строки без наименования или количества
        if not name or not quantity:
            row_num += 1
            continue
        
        try:
            quantity = int(quantity)
        except (ValueError, TypeError):
            quantity = 1
        
        unit = unit_cell.value if unit_cell.value else 'шт'
        code = code_cell.value if code_cell.value else None
        
        # Создаем товар
        item = OrderItem(
            row_number=int(row_number_cell.value),
            name=str(name).strip(),
            quantity=quantity,
            unit=str(unit).strip(),
            code=str(code).strip() if code else None,
            status='pending'
        )
        items.append(item)
        
        row_num += 1
    
    # Добавляем товары к заказу
    order.items = items
    
    # Закрываем файл Excel
    wb.close()
    
    return order


def validate_excel_file(filepath):
    """
    Проверяет корректность Excel файла
    
    Returns:
        tuple: (bool, str) - (валиден ли файл, сообщение об ошибке)
    """
    wb = None
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        
        # Проверяем есть ли данные
        if ws.max_row < 11:
            return False, "Файл не содержит достаточно строк"
        
        # Проверяем наличие заголовка
        header_cell = ws.cell(row=3, column=2)
        if not header_cell.value:
            return False, "Не найден заголовок заказа в строке 3"
        
        return True, "OK"
    
    except Exception as e:
        return False, f"Ошибка чтения файла: {str(e)}"
    
    finally:
        # Всегда закрываем файл
        if wb is not None:
            wb.close()



